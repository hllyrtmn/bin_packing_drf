from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm

from logistics.models import PackageDetail

class SettlementService:
    @staticmethod
    def create_settlement_report(packages):
        wb = Workbook()
        ws = wb.active
        column_names = [
            "No", "Palet Genişlik", "X", "Palet Derinlik", "KOD1", "KOD", "KALAN",
            "Ürün Kodu", "Mod", "Panel Boyu", "Dikey Sıra Adedi","Kat Adedi(Dikey)" ,"Yatay Sıra Adedi",
            "Kat Adedi(Yatay)", "Eşdeğer", "Ara Toplam", "Palet Adet",
            "Metre", "Birim Ağırlık", "Net", "Palet Ağırlık", "Toplam"
        ]
        
        # Kenarlık tanımları
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Başlık satırı
        for col_num, col_name in enumerate(column_names, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border  # Başlıklara kenarlık ekle
        
        # Filtreyi ilk satırda ekle
        ws.auto_filter.ref = f"A1:{get_column_letter(len(column_names))}1"

        row_num = 2  # Veri satırları 2. satırdan başlar
        col_widths = [len(col_name) for col_name in column_names]  # Sütun genişlikleri için başlık uzunluklarını başlangıç değeri olarak alıyoruz
        package_number = 0
        for package in packages:
            package_number += 1
            start_row = row_num
            package_details = PackageDetail.objects.filter(package_id=package.id)
            aa = len(package_details)
            aa = aa
            for package_detail in package_details:
                horizontal_count = 0
                if package.pallet.dimension.width % package_detail.product.dimension.width == 0 and package.pallet.dimension.depth % package_detail.product.dimension.depth == 0:
                    horizontal_count = package.pallet.dimension.width / package_detail.product.dimension.width * package.pallet.dimension.depth / package_detail.product.dimension.depth
                elif package.pallet.dimension.width % package_detail.product.dimension.depth == 0 and package.pallet.dimension.depth % package_detail.product.dimension.width == 0:
                    horizontal_count = package.pallet.dimension.width / package_detail.product.dimension.depth * package.pallet.dimension.depth / package_detail.product.dimension.width
                if horizontal_count == 0:
                    horizontal_count = 1
                values = [
                    package_number,  # "No"
                    package.pallet.dimension.width/10 + 4,  # Palet Genişlik
                    "X",  # X
                    package.pallet.dimension.depth/10 + 2,  # Palet Derinlik
                    f"{package_detail.product.product_type.type}.{package_detail.product.product_type.code}.{int(package_detail.product.dimension.width)}.{int(package_detail.product.dimension.depth)}.{package_detail.count}",  # KOD1
                    f"{package_detail.product.product_type.type}.{package_detail.product.product_type.code}.{int(package_detail.product.dimension.width)}.{int(package_detail.product.dimension.depth)}",  # KOD
                    0,  # Kalan
                    str(f"{package_detail.product.product_type.type}.{int(package_detail.product.dimension.depth)}"),  # Ürün Kodu
                    package_detail.product.product_type.code,  # Mod
                    "",  # Panel Boyu
                    "", # Dikey Sıra Adedi
                    "", # Kat Adedi(Dikey)
                    horizontal_count, #Yatay Sıra Adedi
                    package_detail.count / horizontal_count, # "Kat Adedi(Yatay)"
                    package_detail.count / horizontal_count, # "Eşdeğer",
                    package_detail.count,    #"Ara Toplam",
                    sum(p_detail.count for p_detail in package_details),# Palet Adet
                    int(package_detail.count * package_detail.product.dimension.depth / 1000),  # Metre
                    round(float(package_detail.product.weight_type.std),2),  # Birim Ağırlık
                    round(float(package_detail.product.weight_type.std * package_detail.count),2),  # Net
                    round(float(package.pallet.weight),2),  # Palet Ağırlık
                    round(float(sum(p_detail.product.weight_type.std * p_detail.count for p_detail in package_details) + package.pallet.weight),2)  # Toplam
                ]
                
                # Satır verilerini hücrelere yaz
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Ortalama
                    cell.border = thin_border  # Kenarlık ekle
                    
                    col_widths[col_num - 1] = max(col_widths[col_num - 1], len(str(value)))

                row_num += 1  # Sonraki satır
            if row_num - start_row > 1:
                merge_columns = [1, 2, 3, 4, 17, 21, 22]  # No, Palet Genişlik, X, Palet Derinlik sütunları
                for col in merge_columns:
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=row_num - 1, end_column=col)
                    merge_cell = ws.cell(row=start_row, column=col)
                    merge_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Toplam Satırı Ekle
        total_row = row_num
        total_label_cell = ws.cell(row=total_row, column=15, value="TOPLAM")
        total_label_cell.font = Font(bold=True)
        total_label_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_label_cell.border = thin_border
        
        for col_num in range(16, 23):  # Sayısal değer içeren sütunlar (Palet Adet'ten toplam sütununa kadar)
            col_letter = get_column_letter(col_num)
            sum_formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            total_cell = ws.cell(row=total_row, column=col_num, value=sum_formula)
            total_cell.font = Font(bold=True)
            total_cell.alignment = Alignment(horizontal="center", vertical="center")
            total_cell.border = thin_border  # Kenarlık ekle
            # cell.number_format = "#,##0.00"

        # Sütun genişliklerini yazılara göre ayarla
        for i, width in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = width + 2  # Ekstra boşluk için "+2"

        # Dosyayı kaydet
        wb.save(f"uploads/report_{datetime.now().date()}.xlsx")
        SettlementService.create_pdf_label(packages)
    @staticmethod
    def create_pdf_label(packages):
        # PDF oluşturma
        c = canvas.Canvas("uploads/label_output.pdf", pagesize=A4)
        width, height = A4
        package_number = 1
        for package in packages:
            # Logo ekleme
            logo_width, logo_height = 10 * cm, 3 * cm  # Logo boyutları
            logo_x = (width - logo_width) / 2
            logo_y = height - 5 * cm
            c.drawImage("uploads/bedisa_white_bg.png", logo_x, logo_y, width=logo_width, height=logo_height)

            # Dinamik "NO" bilgisi
            c.setFont("Helvetica-Bold", 45)
            c.drawCentredString(width / 2, height - 8 * cm, f"NO: {package_number}")
            package_number +=1
            # TYPE bilgisi
            c.setFont("Helvetica-Bold", 16)
            c.drawString(3 * cm, height - 11 * cm, "TYPE:")

            # Ürün bilgileri
            c.setFont("Helvetica", 18)
            start_y = height - 13 * cm  # Ürün bilgileri başlangıç konumu
            package_details = PackageDetail.objects.filter(package_id=package.id)
            for package_detail in package_details:
                c.drawString(4 * cm, start_y, f"{package_detail.product.product_type.type}.{package_detail.product.product_type.code}.{int(package_detail.product.dimension.width)}.{int(package_detail.product.dimension.depth)}")
                c.drawString(14 * cm, start_y, f"Q: {sum(p_detail.count for p_detail in package_details)} PCS")
                start_y -= 1 * cm  # Her ürün için satır kaydırma

            # Şirket bilgisi
            c.setFont("Helvetica-Bold", 20)
            c.drawString(3 * cm, height - 20 * cm, f"COMPANY: {package.order.company.company_name}")

            # Ülke bilgisi
            c.setFont("Helvetica-Bold", 30)
            c.drawCentredString(width / 2, height - 23 * cm, "RUSSIA")

            # Yeni sayfa
            c.showPage()

        # PDF'i kaydet
        c.save()